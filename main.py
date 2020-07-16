# This is a GUI application which will show how much tax you need to pay according to a spreadsheet

import tkinter as tk
from openpyxl import Workbook, load_workbook


myworkbook = load_workbook(filename="Tax_table.xlsx", read_only=True)
mysheet = myworkbook.active

# This searches for the header of a spreadsheet
def search_for_header(taxtable="Tax_table.xlsx", requiredheader='Income'):
    foundheader = False
    taxworkbook = load_workbook(filename=taxtable)
    taxsheet = taxworkbook.active
    for eachheader in taxsheet[1]:
        print(eachheader)
        print(eachheader.value)
        print(eachheader.column)
        if eachheader.value == requiredheader:
            foundheader = True
            return eachheader.column

    if foundheader == True:
        pass
    else:
        return 'There is no header'





# This calculates the tax of the user
def calculate_tax(userincome=0, taxtable="Tax_table.xlsx"):
    taxworkbook = load_workbook(filename=taxtable)
    taxsheet = taxworkbook.active
    usertax = 0
    for eachincome in taxsheet.iter_rows(min_col=1, max_col=1, min_row=2):  # The iter_rows returns a tuple with (<cell object>, nul)
        theincome = eachincome[0]  # This is the cell object

        # print(eachincome)
        # print(type(eachincome[0]))
        # print(theincome.value)
        # print(type(theincome.value))

        previncome = taxsheet.cell(row=theincome.row - 1, column=theincome.column).value
        taxrate = taxsheet.cell(row=theincome.row, column=theincome.column + 1).value

        print(previncome)
        # This is for the header being a string
        if previncome == "Income":
            previncome = 0

        # This is for when the user income is higher than the tax bracket
        if theincome.value == "endbracket":
            theincome.value = userincome


        print(previncome)
        print(taxrate)
        print(theincome.value)

        try:
            if userincome >= theincome.value:
                usertax += (theincome.value - previncome) * taxrate
            else:
                usertax += (userincome - previncome) * taxrate
                break

        except:
            print('The comparing income is skipped')

    usertax = int(usertax)  # Rounding down usertax
    return 'Your tax is ' + str(usertax)


def calculate_button(userincome, statuslabel, taxtable='Tax_table.xlsx'):
    allowtocalculate = True
    try:
        userincome = int(userincome)
    except:
        statuslabel['text'] = 'Input is not a number'
        statusmessage.pack()
        allowtocalculate = False

    try:
        if int(userincome) < 0:
            statuslabel['text'] = 'Income has to be higher than zero'
            statuslabel.pack()
            allowtocalculate = False
    except:
        print('the variable userincome is not type of int')

    if allowtocalculate:
        theusertax = calculate_tax(userincome, taxtable)
        statuslabel['text'] = str(theusertax)





### This is the GUI part of the tax calculator ###
HEIGHT = 500
WIDTH = 1000
root = tk.Tk()

frame = tk.Frame(root, height=HEIGHT, width=WIDTH)
frame.pack()

entrylabel = tk.Label(frame, text='Enter income:')
entrylabel.place(relheight=0.1, relwidth=0.1, relx=0.1, rely=0.2)

entrybar = tk.Entry(frame, text='Enter your income here')
entrybar.place(relheight=0.1, relwidth=0.1, relx=0.2, rely=0.2)

statusdisplay = tk.Frame(frame, bg='White')
statusdisplay.place(relheight=0.2, relwidth=0.2, relx=0.1, rely=0.4)

statuslabel = tk.Label(frame, text='Status')
statuslabel.place(relheight=0.1, relwidth=0.1, relx=0.15, rely=0.3)

statusmessage = tk.Label(statusdisplay, text='The message for error/tax', bg='White')
statusmessage.pack()

calculatebutton = tk.Button(frame, text='Calculate', bd=4, command=lambda: calculate_button(entrybar.get(), statusmessage))
calculatebutton.place(relheight=0.1, relwidth=0.2, relx=0.1, rely=0.8)

# This is for the grid display
tableframe = tk.Frame(frame)
tableframe.place(relheight=0.9, relwidth=0.4, relx=0.5, rely=0.2)

taxworkbook = load_workbook('Tax_table.xlsx')
taxsheet = taxworkbook.active
for eachrow in taxsheet.iter_rows():
    for eachcell in eachrow:
        try:
            eachitem = tk.Label(tableframe, text=str(eachcell.value), font=('Helvitica', 25))
            eachitem.grid(row=eachcell.row, column=eachcell.column)
        except:
            print('There is problem in iter for GUI grid table')











root.mainloop()






